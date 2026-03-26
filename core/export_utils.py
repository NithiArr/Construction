"""
Excel export utilities for project data
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from core.models import Project
from finance.models import Expense, Payment, ClientPayment, ExpenseItem
from django.db.models import Sum
from functools import wraps
from django.http import HttpResponseForbidden

def require_export_role(view_func):
    """Ensure only ADMIN or MANAGER can export data."""
    @wraps(view_func)
    def _wrapped(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return HttpResponseForbidden("Authentication required")
        if request.user.role not in ('ADMIN', 'MANAGER'):
            return HttpResponseForbidden("Permission denied. Employees cannot export data.")
        return view_func(request, *args, **kwargs)
    return _wrapped


def apply_header_style(cell):
    """Apply consistent header styling"""
    cell.font = Font(bold=True, color="FFFFFF", size=11)
    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )


def apply_currency_format(ws, col_letter, start_row, end_row):
    """Apply currency format to a column"""
    for row in range(start_row, end_row + 1):
        cell = ws[f'{col_letter}{row}']
        cell.number_format = '₹#,##0.00'


def auto_adjust_column_width(ws):
    """Auto-adjust column widths based on content"""
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width


@login_required
@require_export_role
def export_project_to_excel(request, project_id):
    """Export complete project data to Excel with multiple sheets"""
    try:
        project = Project.objects.get(project_id=project_id, company=request.user.company)
    except Project.DoesNotExist:
        return HttpResponse("Project not found", status=404)
    
    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Get all data
    expenses = Expense.objects.filter(project=project)
    payments = Payment.objects.filter(project=project)
    client_payments = ClientPayment.objects.filter(project=project)
    
    # Calculate KPIs
    material_purchases = expenses.filter(expense_type='Material Purchase')
    regular_expenses = expenses.filter(expense_type='Regular Expense')
    
    total_material = material_purchases.aggregate(Sum('amount'))['amount__sum'] or 0
    total_regular = regular_expenses.aggregate(Sum('amount'))['amount__sum'] or 0
    total_spent = float(total_material) + float(total_regular)
    
    total_vendor_payments = payments.aggregate(Sum('amount'))['amount__sum'] or 0
    outstanding_payables = float(total_material) - float(total_vendor_payments)
    
    total_client_payments = client_payments.aggregate(Sum('amount'))['amount__sum'] or 0
    
    budget_utilized_pct = (total_spent / float(project.budget) * 100) if project.budget > 0 else 0
    
    # ============================================
    # SHEET 1: Project Summary & KPIs
    # ============================================
    ws_summary = wb.create_sheet("Project Summary")
    
    # Project header
    ws_summary['A1'] = f"PROJECT FINANCIAL REPORT"
    ws_summary['A1'].font = Font(bold=True, size=16, color="1F4E78")
    ws_summary.merge_cells('A1:D1')
    
    ws_summary['A2'] = f"Project: {project.name}"
    ws_summary['A2'].font = Font(bold=True, size=14)
    ws_summary.merge_cells('A2:D2')
    
    ws_summary['A3'] = f"Export Date: {datetime.now().strftime('%d-%m-%Y %H:%M')}"
    ws_summary['A3'].font = Font(italic=True)
    ws_summary.merge_cells('A3:D3')
    
    # KPIs section
    row = 5
    ws_summary[f'A{row}'] = "FINANCIAL OVERVIEW"
    ws_summary[f'A{row}'].font = Font(bold=True, size=12, color="FFFFFF")
    ws_summary[f'A{row}'].fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    ws_summary.merge_cells(f'A{row}:D{row}')
    
    row += 1
    kpis = [
        ["Total Budget", f"₹{float(project.budget):,.2f}"],
        ["Total Spent", f"₹{total_spent:,.2f}"],
        ["Budget Utilized", f"{budget_utilized_pct:.1f}%"],
        ["Amount Remaining", f"₹{(float(project.budget) - total_spent):,.2f}"],
        ["", ""],
        ["Material Purchases", f"₹{float(total_material):,.2f}"],
        ["Regular Expenses", f"₹{float(total_regular):,.2f}"],
        ["Vendor Payments Made", f"₹{float(total_vendor_payments):,.2f}"],
        ["Outstanding Payables", f"₹{outstanding_payables:,.2f}"],
        ["", ""],
        ["Client Payments Received", f"₹{float(total_client_payments):,.2f}"],
    ]
    
    for kpi_name, kpi_value in kpis:
        ws_summary[f'A{row}'] = kpi_name
        ws_summary[f'B{row}'] = kpi_value
        if kpi_name:
            ws_summary[f'A{row}'].font = Font(bold=True)
        row += 1
    
    auto_adjust_column_width(ws_summary)
    
    # ============================================
    # SHEET 2: Purchase History (with items)
    # ============================================
    ws_purchases = wb.create_sheet("Purchase History")
    
    headers = ['Date', 'Vendor', 'Invoice No.', 'Category', 'Item Name', 'Quantity', 'Unit', 'Unit Price', 'Total Price', 'Payment Mode']
    for col, header in enumerate(headers, 1):
        cell = ws_purchases.cell(row=1, column=col, value=header)
        apply_header_style(cell)
    
    row = 2
    purchase_total_price = 0.0
    purchase_unit_total = 0.0
    for expense in material_purchases:
        items = ExpenseItem.objects.filter(expense=expense)
        if items.exists():
            for item in items:
                ws_purchases[f'A{row}'] = expense.expense_date.strftime('%d/%m/%Y')
                ws_purchases[f'B{row}'] = expense.vendor.name if expense.vendor else 'N/A'
                ws_purchases[f'C{row}'] = expense.invoice_number or '-'
                ws_purchases[f'D{row}'] = expense.category or 'N/A'
                ws_purchases[f'E{row}'] = item.item_name
                ws_purchases[f'F{row}'] = float(item.quantity)
                ws_purchases[f'G{row}'] = item.measuring_unit
                ws_purchases[f'H{row}'] = float(item.unit_price)
                ws_purchases[f'I{row}'] = float(item.total_price)
                ws_purchases[f'J{row}'] = expense.payment_mode
                purchase_unit_total += float(item.unit_price)
                purchase_total_price += float(item.total_price)
                row += 1
        else:
            # Purchase without items
            ws_purchases[f'A{row}'] = expense.expense_date.strftime('%d/%m/%Y')
            ws_purchases[f'B{row}'] = expense.vendor.name if expense.vendor else 'N/A'
            ws_purchases[f'C{row}'] = expense.invoice_number or '-'
            ws_purchases[f'D{row}'] = expense.category or 'N/A'
            ws_purchases[f'E{row}'] = '-'
            ws_purchases[f'F{row}'] = '-'
            ws_purchases[f'G{row}'] = '-'
            ws_purchases[f'H{row}'] = '-'
            ws_purchases[f'I{row}'] = float(expense.amount)
            ws_purchases[f'J{row}'] = expense.payment_mode
            purchase_total_price += float(expense.amount)
            row += 1
    
    # Apply currency formatting
    if row > 2:
        apply_currency_format(ws_purchases, 'H', 2, row - 1)
        apply_currency_format(ws_purchases, 'I', 2, row - 1)
    
    # Add totals row
    if row > 2:
        ws_purchases[f'H{row}'] = "TOTAL:"
        ws_purchases[f'H{row}'].font = Font(bold=True)
        ws_purchases[f'I{row}'] = purchase_total_price
        ws_purchases[f'I{row}'].font = Font(bold=True)
        ws_purchases[f'I{row}'].number_format = '₹#,##0.00'
    
    ws_purchases.auto_filter.ref = f"A1:J{row}"
    auto_adjust_column_width(ws_purchases)
    
    # ============================================
    # SHEET 3: Regular Expenses
    # ============================================
    ws_expenses = wb.create_sheet("Regular Expenses")
    
    headers = ['Date', 'Category', 'Description', 'Amount', 'Payment Mode']
    for col, header in enumerate(headers, 1):
        cell = ws_expenses.cell(row=1, column=col, value=header)
        apply_header_style(cell)
    
    row = 2
    reg_expense_total = 0.0
    for expense in regular_expenses:
        ws_expenses[f'A{row}'] = expense.expense_date.strftime('%d/%m/%Y')
        ws_expenses[f'B{row}'] = expense.category or 'N/A'
        ws_expenses[f'C{row}'] = expense.description or '-'
        ws_expenses[f'D{row}'] = float(expense.amount)
        ws_expenses[f'E{row}'] = expense.payment_mode
        reg_expense_total += float(expense.amount)
        row += 1
    
    if row > 2:
        apply_currency_format(ws_expenses, 'D', 2, row - 1)
        ws_expenses[f'C{row}'] = "TOTAL:"
        ws_expenses[f'C{row}'].font = Font(bold=True)
        ws_expenses[f'D{row}'] = reg_expense_total
        ws_expenses[f'D{row}'].font = Font(bold=True)
        ws_expenses[f'D{row}'].number_format = '₹#,##0.00'
    
    ws_expenses.auto_filter.ref = f"A1:E{row}"
    auto_adjust_column_width(ws_expenses)
    
    # ============================================
    # SHEET 4: Vendor Payments
    # ============================================
    ws_payments = wb.create_sheet("Vendor Payments")
    
    headers = ['Date', 'Vendor', 'Amount', 'Payment Mode', 'Invoice Reference']
    for col, header in enumerate(headers, 1):
        cell = ws_payments.cell(row=1, column=col, value=header)
        apply_header_style(cell)
    
    row = 2
    vend_pay_total = 0.0
    for payment in payments:
        ws_payments[f'A{row}'] = payment.payment_date.strftime('%d/%m/%Y')
        ws_payments[f'B{row}'] = payment.vendor.name if payment.vendor else 'N/A'
        ws_payments[f'C{row}'] = float(payment.amount)
        ws_payments[f'D{row}'] = payment.payment_mode
        ws_payments[f'E{row}'] = payment.expense.invoice_number if payment.expense else '-'
        vend_pay_total += float(payment.amount)
        row += 1
    
    if row > 2:
        apply_currency_format(ws_payments, 'C', 2, row - 1)
        ws_payments[f'B{row}'] = "TOTAL:"
        ws_payments[f'B{row}'].font = Font(bold=True)
        ws_payments[f'C{row}'] = vend_pay_total
        ws_payments[f'C{row}'].font = Font(bold=True)
        ws_payments[f'C{row}'].number_format = '₹#,##0.00'
    
    ws_payments.auto_filter.ref = f"A1:E{row}"
    auto_adjust_column_width(ws_payments)
    
    # ============================================
    # SHEET 5: Client Payments
    # ============================================
    ws_client = wb.create_sheet("Client Payments")
    
    headers = ['Date', 'Amount', 'Payment Mode', 'Reference Number', 'Remarks']
    for col, header in enumerate(headers, 1):
        cell = ws_client.cell(row=1, column=col, value=header)
        apply_header_style(cell)
    
    row = 2
    client_pay_total = 0.0
    for cp in client_payments:
        ws_client[f'A{row}'] = cp.payment_date.strftime('%d/%m/%Y')
        ws_client[f'B{row}'] = float(cp.amount)
        ws_client[f'C{row}'] = cp.payment_mode
        ws_client[f'D{row}'] = cp.reference_number or '-'
        ws_client[f'E{row}'] = cp.remarks or '-'
        client_pay_total += float(cp.amount)
        row += 1
    
    if row > 2:
        apply_currency_format(ws_client, 'B', 2, row - 1)
        ws_client[f'A{row}'] = "TOTAL:"
        ws_client[f'A{row}'].font = Font(bold=True)
        ws_client[f'B{row}'] = client_pay_total
        ws_client[f'B{row}'].font = Font(bold=True)
        ws_client[f'B{row}'].number_format = '₹#,##0.00'
    
    ws_client.auto_filter.ref = f"A1:E{row}"
    auto_adjust_column_width(ws_client)
    
    # ============================================
    # Generate response
    # ============================================
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"{project.name.replace(' ', '_')}_Export_{datetime.now().strftime('%Y%m%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response


@login_required
@require_export_role
def export_expenses_to_excel(request):
    """Export regular expenses to Excel with optional filtering"""
    
    # Get filter parameters
    project_id = request.GET.get('project')
    date_from_str = request.GET.get('date_from')
    date_to_str = request.GET.get('date_to')
    category = request.GET.get('category')
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Regular Expenses"
    
    # Header
    headers = ['Date', 'Project', 'Category', 'Amount', 'Payment Mode', 'Description']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        apply_header_style(cell)
    
    # Base query
    expenses = Expense.objects.filter(
        company=request.user.company,
        expense_type='Regular Expense'
    ).select_related('project')
    
    # Apply filters & build filename parts
    filename_parts = ["Expenses"]
    
    if project_id:
        expenses = expenses.filter(project_id=project_id)
        try:
            p_name = Project.objects.get(project_id=project_id).name
            filename_parts.append(p_name.replace(' ', '_'))
        except Project.DoesNotExist:
            pass
            
    if category:
        expenses = expenses.filter(category=category)
        filename_parts.append(category.replace(' ', '_'))
        
    if date_from_str:
        try:
            date_from = datetime.strptime(date_from_str, '%Y-%m-%d').date()
            expenses = expenses.filter(expense_date__gte=date_from)
            filename_parts.append(f"from_{date_from_str}")
        except ValueError:
            pass
            
    if date_to_str:
        try:
            date_to = datetime.strptime(date_to_str, '%Y-%m-%d').date()
            expenses = expenses.filter(expense_date__lte=date_to)
            filename_parts.append(f"to_{date_to_str}")
        except ValueError:
            pass
            
    expenses = expenses.order_by('-expense_date')
    
    # Add data rows
    row = 2
    exp_running_total = 0.0
    for expense in expenses:
        ws[f'A{row}'] = expense.expense_date.strftime('%d/%m/%Y')
        ws[f'B{row}'] = expense.project.name if expense.project else 'N/A'
        ws[f'C{row}'] = expense.category or 'N/A'
        ws[f'D{row}'] = float(expense.amount)
        ws[f'E{row}'] = expense.payment_mode
        ws[f'F{row}'] = expense.description or '-'
        exp_running_total += float(expense.amount)
        row += 1
    
    # Apply formatting
    if row > 2:
        apply_currency_format(ws, 'D', 2, row - 1)
        ws[f'C{row}'] = "TOTAL:"
        ws[f'C{row}'].font = Font(bold=True)
        ws[f'D{row}'] = exp_running_total
        ws[f'D{row}'].font = Font(bold=True)
        ws[f'D{row}'].number_format = '₹#,##0.00'
    
    ws.auto_filter.ref = f"A1:F{row}"
    auto_adjust_column_width(ws)
    
    # Generate response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    # Construct filename
    filename_str = "_".join(filename_parts) + f"_Export_{datetime.now().strftime('%Y%m%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename_str}"'
    
    wb.save(response)
    return response


@login_required
@require_export_role
def export_purchases_to_excel(request):
    """Export material purchases with optional filtering"""
    
    # Get filter parameters
    project_id = request.GET.get('project')
    date_from_str = request.GET.get('date_from')
    date_to_str = request.GET.get('date_to')
    vendor_id = request.GET.get('vendor')
    category = request.GET.get('category')
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Material Purchases"
    
    # Header
    headers = ['Date', 'Project', 'Vendor', 'Invoice No.', 'Category', 'Item Name', 'Quantity', 'Unit', 'Unit Price', 'Total Price', 'Payment Mode']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        apply_header_style(cell)
    
    # Base query
    purchases = Expense.objects.filter(
        company=request.user.company,
        expense_type='Material Purchase'
    ).select_related('project', 'vendor')
    
    # Apply filters & build filename parts
    filename_parts = ["Purchases"]
    
    if project_id:
        purchases = purchases.filter(project_id=project_id)
        try:
            p_name = Project.objects.get(project_id=project_id).name
            filename_parts.append(p_name.replace(' ', '_'))
        except Project.DoesNotExist:
            pass

    if vendor_id:
        purchases = purchases.filter(vendor_id=vendor_id)
        try:
            # Assuming vendor is imported, if not use string filter
            pass # Keep filename simple for now
        except:
             pass

    if category:
        purchases = purchases.filter(category=category)
        filename_parts.append(category.replace(' ', '_'))
        
    if date_from_str:
        try:
            date_from = datetime.strptime(date_from_str, '%Y-%m-%d').date()
            purchases = purchases.filter(expense_date__gte=date_from)
            filename_parts.append(f"from_{date_from_str}")
        except ValueError:
            pass
            
    if date_to_str:
        try:
            date_to = datetime.strptime(date_to_str, '%Y-%m-%d').date()
            purchases = purchases.filter(expense_date__lte=date_to)
            filename_parts.append(f"to_{date_to_str}")
        except ValueError:
            pass
            
    purchases = purchases.order_by('-expense_date')
    
    # Add data rows
    row = 2
    pur_running_total = 0.0
    for purchase in purchases:
        items = ExpenseItem.objects.filter(expense=purchase)
        if items.exists():
            for item in items:
                ws[f'A{row}'] = purchase.expense_date.strftime('%d/%m/%Y')
                ws[f'B{row}'] = purchase.project.name if purchase.project else 'N/A'
                ws[f'C{row}'] = purchase.vendor.name if purchase.vendor else 'N/A'
                ws[f'D{row}'] = purchase.invoice_number or '-'
                ws[f'E{row}'] = purchase.category or 'N/A'
                ws[f'F{row}'] = item.item_name
                ws[f'G{row}'] = float(item.quantity)
                ws[f'H{row}'] = item.measuring_unit
                ws[f'I{row}'] = float(item.unit_price)
                ws[f'J{row}'] = float(item.total_price)
                ws[f'K{row}'] = purchase.payment_mode
                pur_running_total += float(item.total_price)
                row += 1
        else:
            # Purchase without items
            ws[f'A{row}'] = purchase.expense_date.strftime('%d/%m/%Y')
            ws[f'B{row}'] = purchase.project.name if purchase.project else 'N/A'
            ws[f'C{row}'] = purchase.vendor.name if purchase.vendor else 'N/A'
            ws[f'D{row}'] = purchase.invoice_number or '-'
            ws[f'E{row}'] = purchase.category or 'N/A'
            ws[f'F{row}'] = '-'
            ws[f'G{row}'] = '-'
            ws[f'H{row}'] = '-'
            ws[f'I{row}'] = '-'
            ws[f'J{row}'] = float(purchase.amount)
            ws[f'K{row}'] = purchase.payment_mode
            pur_running_total += float(purchase.amount)
            row += 1
    
    # Apply formatting
    if row > 2:
        apply_currency_format(ws, 'I', 2, row - 1)
        apply_currency_format(ws, 'J', 2, row - 1)
        ws[f'I{row}'] = "TOTAL:"
        ws[f'I{row}'].font = Font(bold=True)
        ws[f'J{row}'] = pur_running_total
        ws[f'J{row}'].font = Font(bold=True)
        ws[f'J{row}'].number_format = '₹#,##0.00'
    
    ws.auto_filter.ref = f"A1:K{row}"
    auto_adjust_column_width(ws)
    
    # Generate response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    filename_str = "_".join(filename_parts) + f"_Export_{datetime.now().strftime('%Y%m%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename_str}"'
    
    wb.save(response)
    return response


@login_required
@require_export_role
def export_vendor_payments_to_excel(request):
    """Export vendor payments to Excel with filtering"""
    
    # Get filter parameters
    project_id = request.GET.get('project')
    date_from_str = request.GET.get('date_from')
    date_to_str = request.GET.get('date_to')
    vendor_id = request.GET.get('vendor')
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Vendor Payments"
    
    # Header
    headers = ['Date', 'Project', 'Vendor', 'Amount', 'Payment Mode', 'Remarks', 'Invoice Reference']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        apply_header_style(cell)
    
    # Base query
    payments = Payment.objects.filter(
        company=request.user.company
    ).select_related('project', 'vendor', 'expense')
    
    # Apply filters & build filename parts
    filename_parts = ["Vendor_Payments"]
    
    if project_id:
        payments = payments.filter(project_id=project_id)
        try:
            p_name = Project.objects.get(project_id=project_id).name
            filename_parts.append(p_name.replace(' ', '_'))
        except Project.DoesNotExist:
            pass
            
    if vendor_id:
        payments = payments.filter(vendor_id=vendor_id)
        
    if date_from_str:
        try:
            date_from = datetime.strptime(date_from_str, '%Y-%m-%d').date()
            payments = payments.filter(payment_date__gte=date_from)
            filename_parts.append(f"from_{date_from_str}")
        except ValueError:
            pass
            
    if date_to_str:
        try:
            date_to = datetime.strptime(date_to_str, '%Y-%m-%d').date()
            payments = payments.filter(payment_date__lte=date_to)
            filename_parts.append(f"to_{date_to_str}")
        except ValueError:
            pass
            
    payments = payments.order_by('-payment_date')
    
    # Add data rows
    row = 2
    vp_running_total = 0.0
    for payment in payments:
        ws[f'A{row}'] = payment.payment_date.strftime('%d/%m/%Y')
        ws[f'B{row}'] = payment.project.name if payment.project else 'N/A'
        ws[f'C{row}'] = payment.vendor.name if payment.vendor else 'N/A'
        ws[f'D{row}'] = float(payment.amount)
        ws[f'E{row}'] = payment.payment_mode
        ws[f'F{row}'] = payment.expense.invoice_number if payment.expense else '-'
        vp_running_total += float(payment.amount)
        row += 1
    
    # Apply formatting
    if row > 2:
        apply_currency_format(ws, 'D', 2, row - 1)
        ws[f'C{row}'] = "TOTAL:"
        ws[f'C{row}'].font = Font(bold=True)
        ws[f'D{row}'] = vp_running_total
        ws[f'D{row}'].font = Font(bold=True)
        ws[f'D{row}'].number_format = '₹#,##0.00'
    
    ws.auto_filter.ref = f"A1:F{row}"
    auto_adjust_column_width(ws)
    
    # Generate response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    filename_str = "_".join(filename_parts) + f"_Export_{datetime.now().strftime('%Y%m%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename_str}"'
    
    wb.save(response)
    return response


@login_required
@require_export_role
def export_cash_balance_excel(request):
    """
    Export Cash Balance report to Excel.
    Single project  → Sheet 1: vertical summary | Sheet 2-5: one breakdown sheet each
    Multi / All     → Sheet 1: project-per-row summary | Sheets 2-5: one per breakdown type
    Credit purchases are always highlighted in red.
    """
    from finance.models import Expense, Payment, ClientPayment
    from core.models import Project
    from django.db.models import Sum
    from datetime import datetime

    # Accept project_ids (new multi) or project_id (legacy single)
    project_ids_str = request.GET.get('project_ids') or request.GET.get('project_id')
    from_date_str = request.GET.get('from_date')
    to_date_str   = request.GET.get('to_date')

    if not from_date_str or not to_date_str:
        return HttpResponse('Date range required', status=400)

    from_date = datetime.fromisoformat(from_date_str).date()
    to_date   = datetime.fromisoformat(to_date_str).date()

    # ── Resolve which projects are selected ────────────────────────────────────
    all_projects_qs = Project.objects.filter(company=request.user.company).order_by('name')
    is_multi = False  # single project mode by default

    if not project_ids_str or project_ids_str == 'all':
        selected_projects = list(all_projects_qs)
        is_multi = True
        filename_label = 'All_Projects'
    else:
        id_list = [pid.strip() for pid in project_ids_str.split(',') if pid.strip()]
        selected_projects = list(all_projects_qs.filter(project_id__in=id_list))
        if len(selected_projects) == 1:
            filename_label = selected_projects[0].name.replace(' ', '_')
        else:
            is_multi = True
            filename_label = 'Multi_Projects'

    # ── Colour / style constants ───────────────────────────────────────────────
    RED_FONT     = Font(color='FF0000')
    RED_FILL     = PatternFill(start_color='FFF0F0', end_color='FFF0F0', fill_type='solid')
    BOLD         = Font(bold=True)
    SECTION_FILL = PatternFill(start_color='2E4057', end_color='2E4057', fill_type='solid')
    SECTION_FONT = Font(bold=True, color='FFFFFF', size=11)
    TOTAL_FILL   = PatternFill(start_color='D6EAF8', end_color='D6EAF8', fill_type='solid')
    IN_FONT      = Font(color='1E8449', bold=True)
    OUT_FONT     = Font(color='C0392B', bold=True)
    IN_FILL      = PatternFill(start_color='E9F7EF', end_color='E9F7EF', fill_type='solid')
    OUT_FILL     = PatternFill(start_color='FDEDEC', end_color='FDEDEC', fill_type='solid')
    THIN_BORDER  = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    INR_FMT = '₹#,##0.00'

    def hdr(cell, text=None):
        if text is not None:
            cell.value = text
        apply_header_style(cell)

    def write_breakdown_sheet(wb, sheet_title, col_headers, rows_data, amount_col_0idx,
                               red_row_fn=None):
        """Write a single standalone breakdown sheet."""
        ws = wb.create_sheet(sheet_title)
        nc = len(col_headers)
        # Section banner
        c = ws.cell(row=1, column=1, value=sheet_title.upper())
        c.font = SECTION_FONT; c.fill = SECTION_FILL
        c.alignment = Alignment(horizontal='left', vertical='center')
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=nc)
        ws.row_dimensions[1].height = 20
        # Headers
        for ci, h in enumerate(col_headers, 1):
            hdr(ws.cell(row=2, column=ci), h)
        # Data
        r = 3
        for row_t in rows_data:
            is_red = red_row_fn and red_row_fn(row_t)
            for ci, val in enumerate(row_t, 1):
                c = ws.cell(row=r, column=ci, value=val)
                c.border = THIN_BORDER
                if is_red:
                    c.font = RED_FONT; c.fill = RED_FILL
                if (ci - 1) == amount_col_0idx:
                    c.number_format = INR_FMT
                    c.alignment = Alignment(horizontal='right')
            r += 1
        # Totals row
        if rows_data:
            total_col = amount_col_0idx + 1
            total_val = sum(float(rt[amount_col_0idx]) for rt in rows_data)
            for ci in range(1, nc + 1):
                c = ws.cell(row=r, column=ci)
                c.border = THIN_BORDER; c.fill = TOTAL_FILL
            ws.cell(row=r, column=total_col - 1, value='TOTAL').font = BOLD
            tc = ws.cell(row=r, column=total_col, value=total_val)
            tc.font = BOLD; tc.number_format = INR_FMT
            tc.alignment = Alignment(horizontal='right')
        ws.auto_filter.ref = f"A2:{get_column_letter(nc)}{r}"
        auto_adjust_column_width(ws)
        return ws

    # ── Per-project balance calculation helper ────────────────────────────────
    def calc_balance(project_obj):
        eq = Expense.objects.filter(company=request.user.company, project=project_obj)
        pq = Payment.objects.filter(company=request.user.company, project=project_obj)
        cq = ClientPayment.objects.filter(company=request.user.company, project=project_obj)

        op_in  = float(cq.filter(payment_date__lt=from_date).aggregate(s=Sum('amount'))['s'] or 0)
        op_eo  = float(eq.filter(expense_date__lt=from_date).exclude(payment_mode='CREDIT').aggregate(s=Sum('amount'))['s'] or 0)
        op_po  = float(pq.filter(payment_date__lt=from_date).aggregate(s=Sum('amount'))['s'] or 0)
        opening = op_in - op_eo - op_po

        p_exp = eq.filter(expense_date__range=[from_date, to_date])
        p_pay = pq.filter(payment_date__range=[from_date, to_date])
        p_cp  = cq.filter(payment_date__range=[from_date, to_date])

        receipts    = float(p_cp.aggregate(s=Sum('amount'))['s'] or 0)
        vend_pmts   = float(p_pay.aggregate(s=Sum('amount'))['s'] or 0)
        exp_credit  = float(p_exp.filter(payment_mode='CREDIT').aggregate(s=Sum('amount'))['s'] or 0)
        exp_cash    = float(p_exp.filter(payment_mode='CASH').aggregate(s=Sum('amount'))['s'] or 0)
        exp_bankupi = float(p_exp.filter(payment_mode__in=['BANK', 'UPI']).aggregate(s=Sum('amount'))['s'] or 0)
        outflow     = exp_cash + exp_bankupi + vend_pmts
        closing     = opening + receipts - outflow
        return {
            'opening': opening, 'receipts': receipts, 'vend_pmts': vend_pmts,
            'exp_credit': exp_credit, 'exp_cash': exp_cash, 'exp_bankupi': exp_bankupi,
            'closing': closing,
            'p_exp': p_exp, 'p_pay': p_pay, 'p_cp': p_cp
        }

    # ── Compute all project data upfront ──────────────────────────────────────
    project_data = [(proj, calc_balance(proj)) for proj in selected_projects]

    # ════════════════════════════════════════════════════════════════════════════
    # WORKBOOK
    # ════════════════════════════════════════════════════════════════════════════
    wb = Workbook()
    ws1 = wb.active
    ws1.title = 'Summary'

    # ── Common title block ────────────────────────────────────────────────────
    period_str = f'{from_date.strftime("%d-%m-%Y")} to {to_date.strftime("%d-%m-%Y")}'
    ws1['A1'] = 'DAILY CASH BALANCE REPORT'
    ws1['A1'].font = Font(bold=True, size=16, color='1F4E78')
    ws1['A2'] = f'Period: {period_str}'
    ws1['A2'].font = Font(bold=True, size=11)
    ws1['A3'] = f'Exported: {datetime.now().strftime("%d-%m-%Y %H:%M")}'
    ws1['A3'].font = Font(italic=True, color='888888')

    # ════════════════════════════════════════════════════════════════════════════
    # SINGLE PROJECT – vertical summary (original format)
    # ════════════════════════════════════════════════════════════════════════════
    if not is_multi:
        proj, bd = project_data[0]
        ws1['A2'] = f'Project: {proj.name}   |   Period: {period_str}'
        ws1['A2'].font = Font(bold=True, size=11)
        ws1.merge_cells('A1:C1'); ws1.merge_cells('A2:C2'); ws1.merge_cells('A3:C3')

        summary_rows = [
            ('Opening Balance',             bd['opening'],    'neutral'),
            ('+ Client Receipts (Period)',   bd['receipts'],   'in'),
            ('– Vendor Payments (Period)',   bd['vend_pmts'],  'out'),
            ('– Expenses Credit (Period)',   bd['exp_credit'], 'out'),
            ('– Expenses Cash (Period)',     bd['exp_cash'],   'out'),
            ('– Expenses UPI/Bank (Period)', bd['exp_bankupi'],'out'),
            ('Closing Balance',             bd['closing'],    'closing'),
        ]
        r = 5
        for label, value, kind in summary_rows:
            lc = ws1.cell(row=r, column=1, value=label)
            vc = ws1.cell(row=r, column=3, value=value)
            vc.number_format = INR_FMT
            vc.alignment = Alignment(horizontal='right')
            if kind == 'in':
                lc.font = IN_FONT; vc.font = IN_FONT
                ws1.cell(row=r, column=1).fill = IN_FILL
                ws1.cell(row=r, column=3).fill = IN_FILL
            elif kind == 'out':
                lc.font = OUT_FONT; vc.font = OUT_FONT
                ws1.cell(row=r, column=1).fill = OUT_FILL
                ws1.cell(row=r, column=3).fill = OUT_FILL
            elif kind == 'closing':
                cl_color = '1E8449' if bd['closing'] >= 0 else 'C0392B'
                lc.font = Font(bold=True, size=12, color=cl_color)
                vc.font = Font(bold=True, size=12, color=cl_color)
                ws1.row_dimensions[r].height = 22
            else:
                lc.font = BOLD; vc.font = BOLD
            for c in [1, 3]:
                ws1.cell(row=r, column=c).border = THIN_BORDER
            r += 1
        ws1.column_dimensions['A'].width = 34
        ws1.column_dimensions['B'].width = 4
        ws1.column_dimensions['C'].width = 18

        p_exp = bd['p_exp']
        p_pay = bd['p_pay']
        p_cp  = bd['p_cp']

    # ════════════════════════════════════════════════════════════════════════════
    # MULTI / ALL PROJECT – row-per-project summary
    # ════════════════════════════════════════════════════════════════════════════
    else:
        ws1.merge_cells('A1:I1'); ws1.merge_cells('A2:I2'); ws1.merge_cells('A3:I3')
        # Column headers
        COL_HDRS = [
            'Project',
            'Opening Balance',
            '+ Client Receipts',
            '– Vendor Payments',
            '– Expenses Credit',
            '– Expenses Cash',
            '– Expenses UPI/Bank',
            'Closing Balance',
        ]
        r = 5
        for ci, h in enumerate(COL_HDRS, 1):
            c = ws1.cell(row=r, column=ci, value=h)
            apply_header_style(c)
        r += 1

        # Per-project rows
        totals = [0.0] * 7  # opening, receipts, vend, credit, cash, bankupi, closing
        for proj, bd in project_data:
            vals = [bd['opening'], bd['receipts'], bd['vend_pmts'],
                    bd['exp_credit'], bd['exp_cash'], bd['exp_bankupi'], bd['closing']]
            # Project name cell
            nc_cell = ws1.cell(row=r, column=1, value=proj.name)
            nc_cell.font = BOLD; nc_cell.border = THIN_BORDER
            for ci, val in enumerate(vals, 2):
                c = ws1.cell(row=r, column=ci, value=val)
                c.number_format = INR_FMT
                c.border = THIN_BORDER
                c.alignment = Alignment(horizontal='right')
                # Colour inflow / outflow / closing
                if ci == 3:   # receipts
                    c.font = IN_FONT; c.fill = IN_FILL
                elif 4 <= ci <= 7:  # outflows
                    c.font = OUT_FONT; c.fill = OUT_FILL
                elif ci == 8:  # closing
                    cl_color = '1E8449' if val >= 0 else 'C0392B'
                    c.font = Font(bold=True, color=cl_color)
                totals[ci - 2] += val
            r += 1

        # Totals row
        ws1.cell(row=r, column=1, value='TOTAL').font = Font(bold=True, size=11)
        ws1.cell(row=r, column=1).fill = TOTAL_FILL
        ws1.cell(row=r, column=1).border = THIN_BORDER
        for ci, tot in enumerate(totals, 2):
            c = ws1.cell(row=r, column=ci, value=tot)
            c.number_format = INR_FMT; c.fill = TOTAL_FILL
            c.border = THIN_BORDER; c.alignment = Alignment(horizontal='right')
            # Match column colour coding: receipts=green, outflows=red, closing=green/red
            if ci == 3:   # + Client Receipts
                c.font = Font(bold=True, color='1E8449')
            elif 4 <= ci <= 7:  # outflows (vendor pmts, exp credit/cash/bankupi)
                c.font = Font(bold=True, color='C0392B')
            elif ci == 8:  # Closing Balance
                cl_color = '1E8449' if tot >= 0 else 'C0392B'
                c.font = Font(bold=True, size=11, color=cl_color)
            else:  # Opening Balance
                c.font = BOLD

        auto_adjust_column_width(ws1)

        # Aggregate period queries across all selected projects
        proj_ids = [str(p.project_id) for p, _ in project_data]
        p_exp = Expense.objects.filter(company=request.user.company, project_id__in=proj_ids,
                                        expense_date__range=[from_date, to_date])
        p_pay = Payment.objects.filter(company=request.user.company, project_id__in=proj_ids,
                                        payment_date__range=[from_date, to_date])
        p_cp  = ClientPayment.objects.filter(company=request.user.company, project_id__in=proj_ids,
                                              payment_date__range=[from_date, to_date])

    # ════════════════════════════════════════════════════════════════════════════
    # BREAKDOWN SHEETS (4 separate sheets)
    # ════════════════════════════════════════════════════════════════════════════

    # ── Sheet: Regular Expenses ───────────────────────────────────────────────
    reg_expenses = p_exp.filter(expense_type='Regular Expense').order_by('-expense_date').select_related('project')
    reg_rows = []
    for e in reg_expenses:
        first_item = e.items.first()
        sub = first_item.item_name if first_item else '-'
        reg_rows.append((
            e.expense_date.strftime('%d/%m/%Y'), e.project.name,
            e.category or '-', sub, float(e.amount), e.payment_mode, e.description or '-',
        ))
    write_breakdown_sheet(wb, 'Regular Expenses',
        ['Date', 'Project', 'Category', 'Sub-Category', 'Amount', 'Payment Mode', 'Description'],
        reg_rows, amount_col_0idx=4)

    # ── Sheet: Material Purchases ─────────────────────────────────────────────
    mat_purchases = p_exp.filter(expense_type='Material Purchase').order_by('-expense_date').select_related('project', 'vendor')
    mat_rows = []
    for e in mat_purchases:
        mat_rows.append((
            e.expense_date.strftime('%d/%m/%Y'), e.project.name,
            e.vendor.name if e.vendor else '-', e.invoice_number or '-',
            e.category or '-', float(e.amount), e.payment_mode,
        ))
    write_breakdown_sheet(wb, 'Material Purchases',
        ['Date', 'Project', 'Vendor', 'Invoice No.', 'Category', 'Amount', 'Payment Mode'],
        mat_rows, amount_col_0idx=5,
        red_row_fn=lambda rt: rt[6] == 'CREDIT')

    # ── Sheet: Vendor Payments ────────────────────────────────────────────────
    vend_pays = p_pay.order_by('-payment_date').select_related('project', 'vendor', 'expense')
    vp_rows = []
    for p in vend_pays:
        vp_rows.append((
            p.payment_date.strftime('%d/%m/%Y'), p.project.name,
            p.vendor.name if p.vendor else '-', float(p.amount), p.payment_mode,
            p.expense.invoice_number if p.expense else '-',
        ))
    write_breakdown_sheet(wb, 'Vendor Payments',
        ['Date', 'Project', 'Vendor', 'Amount', 'Payment Mode', 'Invoice Ref'],
        vp_rows, amount_col_0idx=3)

    # ── Sheet: Client Payments ────────────────────────────────────────────────
    client_pays = p_cp.order_by('-payment_date').select_related('project')
    cp_rows2 = []
    for cp in client_pays:
        cp_rows2.append((
            cp.payment_date.strftime('%d/%m/%Y'), cp.project.name,
            float(cp.amount), cp.payment_mode, cp.reference_number or '-', cp.remarks or '-',
        ))
    write_breakdown_sheet(wb, 'Client Payments',
        ['Date', 'Project', 'Amount', 'Payment Mode', 'Reference No.', 'Remarks'],
        cp_rows2, amount_col_0idx=2)

    # ── Build response ────────────────────────────────────────────────────────
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f'CashBalance_{filename_label}_{from_date_str}_to_{to_date_str}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response



@login_required
@require_export_role
def export_client_payments_to_excel(request):
    """Export client payments to Excel with filtering"""
    
    # Get filter parameters
    project_id = request.GET.get('project')
    date_from_str = request.GET.get('date_from')
    date_to_str = request.GET.get('date_to')
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Client Payments"
    
    # Header
    headers = ['Date', 'Project', 'Amount', 'Payment Mode', 'Reference Number', 'Remarks']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        apply_header_style(cell)
    
    # Base query
    payments = ClientPayment.objects.filter(
        company=request.user.company
    ).select_related('project')
    
    # Apply filters & build filename parts
    filename_parts = ["Client_Payments"]
    
    if project_id:
        payments = payments.filter(project_id=project_id)
        try:
            p_name = Project.objects.get(project_id=project_id).name
            filename_parts.append(p_name.replace(' ', '_'))
        except Project.DoesNotExist:
            pass
            
    if date_from_str:
        try:
            date_from = datetime.strptime(date_from_str, '%Y-%m-%d').date()
            payments = payments.filter(payment_date__gte=date_from)
            filename_parts.append(f"from_{date_from_str}")
        except ValueError:
            pass
            
    if date_to_str:
        try:
            date_to = datetime.strptime(date_to_str, '%Y-%m-%d').date()
            payments = payments.filter(payment_date__lte=date_to)
            filename_parts.append(f"to_{date_to_str}")
        except ValueError:
            pass
            
    payments = payments.order_by('-payment_date')
    
    # Add data rows
    row = 2
    cp_running_total = 0.0
    for payment in payments:
        ws[f'A{row}'] = payment.payment_date.strftime('%d/%m/%Y')
        ws[f'B{row}'] = payment.project.name if payment.project else 'N/A'
        ws[f'C{row}'] = float(payment.amount)
        ws[f'D{row}'] = payment.payment_mode
        ws[f'E{row}'] = payment.reference_number or '-'
        ws[f'F{row}'] = payment.remarks or '-'
        cp_running_total += float(payment.amount)
        row += 1
    
    # Apply formatting
    if row > 2:
        apply_currency_format(ws, 'C', 2, row - 1)
        ws[f'B{row}'] = "TOTAL:"
        ws[f'B{row}'].font = Font(bold=True)
        ws[f'C{row}'] = cp_running_total
        ws[f'C{row}'].font = Font(bold=True)
        ws[f'C{row}'].number_format = '₹#,##0.00'
    
    ws.auto_filter.ref = f"A1:F{row}"
    auto_adjust_column_width(ws)
    
    # Generate response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    filename_str = "_".join(filename_parts) + f"_Export_{datetime.now().strftime('%Y%m%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename_str}"'
    
    wb.save(response)
    return response
